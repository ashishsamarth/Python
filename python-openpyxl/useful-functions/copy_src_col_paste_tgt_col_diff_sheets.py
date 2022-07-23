from openpyxl.utils import column_index_from_string
# Method to copy data from source and paste it to target of different worksheets in same workbook
# Arguments to this method are: - Source worksheet name, Column Letter of Source, Target worksheet name
# and Column letter of Target
def copy_src_col_paste_tgt_col_diff_sheets(self, _header_row_num, _src_ws_name, _src_col_name, _tgt_ws_name,
                                           _tgt_col_name):
    # Fetch the column index from column name using 'ref_col_name_letter_map' method
    # Since we are dealing with multiple worksheets in this case
    # We must activate each worksheet before pulling column indexes
    self.active_ws(_src_ws_name)
    _src_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_src_col_name])
    self.active_ws(_tgt_ws_name)
    _tgt_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_tgt_col_name])
    # Enumerate over values (by rows) of the worksheet
    for _xl_row_idx, _row_val in enumerate(self.my_base_wb[_src_ws_name].iter_rows(values_only=True), start=1):
        # Write the contents of cells to target column from source column excluding header row
        self.my_base_wb[_tgt_ws_name].cell(row=_xl_row_idx + _header_row_num, column=_tgt_col_idx).value = \
            self.my_base_wb[
                _src_ws_name].cell(row=_xl_row_idx + _header_row_num, column=_src_col_idx).value
    # Save workbook
    self.save_wb()