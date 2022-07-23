from openpyxl.utils import get_column_letter
# Method to automatically adjust column width based on max cell value in that columns
# This method has no argument
def adjust_col_width_in_active_ws(self):
    # Enumerate of the columns in the worksheet and start the index from 1
    # Start value is '1' since 'get_column_letter' method does not allow an index of 0
    for _col_num, _col in enumerate(self.my_base_active_ws.columns, start=1):
        # Get the column letter based on the index
        _col_letter = get_column_letter(_col_num)
        # get the length of maximum cell value (width-wise)
        _col_length = max(len(str(cell.value or "")) for cell in _col)
        # create a new variable adjusted width with a new formulated value
        adjusted_width = (_col_length + 2) * 0.95
        if adjusted_width > 30:
            # use the column dimensions parameter and set the width with a new value
            self.my_base_active_ws.column_dimensions[(str(_col_letter))].width = 55
        else:
            # use the column dimensions parameter and set the width with a new value
            self.my_base_active_ws.column_dimensions[(str(_col_letter))].width = adjusted_width
        self.save_wb()
