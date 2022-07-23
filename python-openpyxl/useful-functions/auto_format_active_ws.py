from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
# Method to auto format the rows and columns in the worksheet
# Argument to this method are: - Worksheet name and Row number of header
def auto_format_active_ws(self, _ws_name, _header_row_num):
    # Activate the worksheet to be formatted
    self.active_ws(_ws_name)
    # Iterate over the number of columns in the worksheet
    for _col in range(1, self.my_base_active_ws_max_col + 1):
        # Iterate over the number of rows in the worksheet
        for _row in range(_header_row_num, self.my_base_active_ws_max_row + 1):
            # Set font, border and alignment for non header cells
            # Identify cell range for non-header row
            filled_cells = self.my_base_active_ws.cell(_row, _col)
            # set the alignment of the non-header row
            filled_cells.alignment = Alignment(wrap_text=False, vertical='top', horizontal='left')
            # set the font of the non-header row
            filled_cells.font = Font(name=self.cell_font_name, size=self.cell_font_size)
            # set the border for non-header row
            filled_cells.border = self.thin_border
        # Set font, border and alignment for header cells
        # Identify cell range for header row
        _cell_header = self.my_base_active_ws.cell(_header_row_num, _col)
        # set the alignment of the header row
        _cell_header.alignment = Alignment(wrap_text=False, vertical='top', horizontal='center')
        # set the font of the header row
        _cell_header.font = Font(name=self.header_font_name, size=self.header_font_size, bold=True)
        # set the color fill for header row
        _cell_header.fill = self.header_fill
        # set the border for header row
        _cell_header.border = self.thick_border
    self.save_wb()
