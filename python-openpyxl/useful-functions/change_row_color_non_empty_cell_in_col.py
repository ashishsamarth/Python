from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
# Method to change row color based on a non-empty cell in particular column
# Arguments to this method: - Row number for header row, Column name, and user selected fill color
def change_row_color_non_empty_cell_in_col(self, _header_row_num, _col_name, _fill_color):
    # Fetch the column index from column name using 'ref_col_name_letter_map' method
    _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
    # Iterate over the values of the column and get row indexes of non-empty cells excluding header row
    # The Excel columns start with 1, however when iterating, the tuples start with index 0
    _row_idx = [_xl_row_idx for _xl_row_idx, _row_val in
                enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1) if _row_val[_col_idx - 1] if
                _xl_row_idx != _header_row_num]
    # define the color pattern and using the class variable match the color hex code
    _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                            fill_type='solid')
    # Iterate over the list of row indexes
    for _row_num in _row_idx:
        # Iterate over the columns of the row
        for _ in range(1, self.my_base_active_ws_max_col + 1):
            # Identify the cells based on row and column indexes
            filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_)
            # apply color
            filled_cells.fill = _my_color
    self.save_wb()