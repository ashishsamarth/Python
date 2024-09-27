from openpyxl.styles import PatternFill
# Method to change row color based on an exact match of cell value in particular column
# Arguments to this method: - Row number for header row, Column name, value of cell, user selected fill color
def change_row_color_by_cell_value_exact_match(self, _header_row_num, _col_name, _cell_value, _fill_color):
    # Get the list of indexes where near match of searched value if found excluding the header row
    _row_idx = self.get_row_idx_lst_based_on_search_val_specific_col_exact_match(_header_row_num, _col_name,
                                                                                 _cell_value)
    # Set the color fill based on user selected color
    _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                            fill_type='solid')
    # Iterate over the list of row indexes identified earlier
    for _row_num in _row_idx:
        # Iterate over each column of the row
        for _ in range(1, self.my_base_active_ws_max_col + 1):
            # Identify cell range with matching criteria
            filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_)
            # set the color fill for row
            filled_cells.fill = _my_color
    self.save_wb()