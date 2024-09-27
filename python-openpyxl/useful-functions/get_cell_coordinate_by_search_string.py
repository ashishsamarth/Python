from openpyxl.utils import get_column_letter
# Method to get list of cell coordinate by a search value
# Argument to this method is:- search string
def get_cell_coordinate_by_search_string(self, _search_val):
    # Create an empty list
    _match_row_idx_lst = []
    # Enumerate over the values (in rows) with starting index of 1
    for _xl_row_idx, _row_val in enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1):
        # If search string is in the cell value
        # Ignore Case while checking membership
        if str(_search_val).casefold() in str(_row_val).casefold():
            # Iterate over the length of cell values to get the index
            for _idx in range(len(_row_val)):
                if str(_search_val).casefold() == str(_row_val[_idx]).casefold():
                    # get the index position and concatenate the Column Letter with row index
                    # Append the identified cell coordinate to empty list
                    _match_row_idx_lst.append(str(get_column_letter(_idx + 1) + str(_xl_row_idx)))
    # Return type is a list
    return _match_row_idx_lst