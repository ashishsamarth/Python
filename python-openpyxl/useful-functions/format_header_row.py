from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
# Method to format header row with user defined, cell color and font color
# Arguments to this method are: - Row number of the header row, fill color for the cells in header row
# Font color in the cells
def format_header_row(self, _header_row_num, _cell_color, _font_color):
    # Iterate over the cells in the header row to the max column size
    for _col_idx in range(1, self.my_base_active_ws_max_col + 1):
        # Set the header cells for easy style formatting
        header_cells = self.my_base_active_ws.cell(row=_header_row_num, column=_col_idx)
        # Set header cell alignment
        header_cells.alignment = Alignment(wrap_text=False, vertical='top', horizontal='center')
        # Set header cell border
        header_cells.border = self.thick_border
        # Fill the cells with user selected color from the color map
        header_cells.fill = PatternFill(start_color=self._my_color_map[_cell_color],
                                        end_color=self._my_color_map[_cell_color],
                                        fill_type="solid")
        # Change the color of the font in cells with user selected color
        header_cells.font = Font(bold=True, size=self.header_font_size,
                                 name=self.header_font_name,
                                 color=self._my_color_map[_font_color])
    self.save_wb()