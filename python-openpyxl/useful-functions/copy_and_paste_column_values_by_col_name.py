from openpyxl.utils import column_index_from_string
# Method to copy and paste data from one column to another, excluding header
# Arguments to this method are: - name of source column, name of target column & row number for header
def copy_and_paste_column_values_by_col_name(self, _src_col_name, _tgt_col_name, _header_row_num):
    # Fetch the column index from column name using 'ref_col_name_letter_map' method
    _src_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_src_col_name])
    _tgt_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_tgt_col_name])
    # Enumerate over values (by rows) of the worksheet
    for _xl_row_idx, _row_val in enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1):
        # Exclude the header row
        if _xl_row_idx != _header_row_num:
            # Write the contents of cells to target column from source column
            self.my_base_active_ws.cell(row=_xl_row_idx,
                                        column=_tgt_col_idx).value = self.my_base_active_ws.cell(
                row=_xl_row_idx, column=_src_col_idx).value
    # Save workbook
    self.save_wb()