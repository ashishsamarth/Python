from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

_my_color_map = {'blue': '0000CCFF',
                     'black': '00000000',
                     'light_green': '00CCFFCC',
                     'navy_blue': '00003366',
                     'green': '00008000',
                     'gray': '00969696',
                     'orange': '00FF9900',
                     'red': '00FF0000',
                     'turquoise': '0000FFFF',
                     'olive_green': '00808000',
                     'white': '00FFFFFF',
                     'yellow': '00FFFF00'}
                     
# Method to change row color based on an exact match of cell value in particular column
# Arguments to this method: - Row number for header row, Column name, value of cell, user selected fill color
def change_cell_color_by_cell_value_exact_match(self, _header_row_num, _col_name, _cell_value, _fill_color):
    # Fetch the column index from column name using 'ref_col_name_letter_map' method
    _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
    # Get the list of indexes where near match of searched value if found excluding the header row
    _row_idx = self.get_row_idx_lst_based_on_search_val_specific_col_exact_match(_header_row_num, _col_name,
                                                                                 _cell_value)
    # Set the color fill based on user selected color
    _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                            fill_type='solid')
    # Iterate over the list of row indexes identified earlier
    for _row_num in _row_idx:
        # Identify cell range with matching criteria
        # Note: The updated index value is to be used since we are writing to the cell
        filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_col_idx)
        # set the color fill for row
        filled_cells.fill = _my_color
    self.save_wb()