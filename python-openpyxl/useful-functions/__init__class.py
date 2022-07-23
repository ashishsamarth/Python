# Import openpyxl
# Note: openpyxl package provides both read and write capabilities to excel
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows


# Class definitions should use CamelCase convention based on pep-8 guidelines
class CustomOpenpyxl:
    # Class Variables
    # Color map as class variable to extend accessibility to all methods
    _my_color_map = {'blue': '0000CCFF',
                     'black': '00000000',
                     'light_green': '00CCFFCC',
                     'navy_blue': '00003366',
                     'green': '00008000',
                     'gray': '00969696',
                     'orange': '00FF9900',
                     'red': '00FF0000',
                     'turquoise': '0000FFFF',
                     'olive_green': '00808000',
                     'white': '00FFFFFF',
                     'yellow': '00FFFF00'}

    # Define Color and border parameters as class variables to extend accessibility to all methods
    header_fill = PatternFill(start_color='CD5C5C', end_color='CD5C5C', fill_type="solid")
    even_row_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type="solid")
    odd_row_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                          bottom=Side(style='thick'))

    # Set Header font and font type as class variables to extend accessibility to all methods
    header_font_name = 'Amasis MT Pro'
    header_font_size = 9
    # Set value font and font type as class variables to extend accessibility to all methods
    cell_font_name = 'Cambria'
    cell_font_size = 8

    # Initialize the class with filename as only argument
    def __init__(self, _my_file_name):
        assert _my_file_name.split('.')[-1] in ['xls', 'xlsx'], 'Input file does not have valid excel extension'
        self.my_filename = _my_file_name
        self.my_base_wb = openpyxl.load_workbook(self.my_filename, read_only=False)
        # following line will get the names of worksheets in the workbook
        self.ws_names_in_my_base_wb = self.my_base_wb.sheetnames
        # following line will get the number of worksheets in the workbook
        self.num_ws_in_my_base_wb = len(self.my_base_wb.sheetnames)
        # following line will set the last worksheet in the workbook as active
        self.my_base_active_ws = self.my_base_wb.active
        # following line will get the maximum row number in active worksheet
        self.my_base_active_ws_max_row = self.my_base_wb.active.max_row
        # following line will get the maximum column number in active worksheet
        self.my_base_active_ws_max_col = self.my_base_active_ws.max_column
        # following line will get the column names (as list) in active worksheet
        # Assumption is the column names are in 1st row
        self.my_base_active_ws_col_titles = [_col_name[0].value for _col_name in self.my_base_active_ws.iter_cols()]
