# Note: openpyxl package provides both read and write capabilities to excel
from curses import start_color
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows


# Class definitions should use CamelCase convention based on pep-8 guidelines
class CustomOpenpyxl:
    # Class Variables
    # Color map as class variable to extend accessibility to all methods
    _my_color_map = {'blue': '0000CCFF',
                     'black': '00000000',
                     'light_green': '00CCFFCC',
                     'navy_blue': '00003366',
                     'green': '00008000',
                     'gray': '00969696',
                     'orange': '00FF9900',
                     'red': '00FF0000',
                     'turquoise': '0000FFFF',
                     'olive_green': '00808000',
                     'white': '00FFFFFF',
                     'yellow': '00FFFF00'}

    # Set Header font and font type as class variables to extend accessibility to all methods
    header_font_name = 'Amasis MT Pro'
    header_font_size = 9
    # Set value font and font type as class variables to extend accessibility to all methods
    cell_font_name = 'Cambria'
    cell_font_size = 8

    # Define Color and border parameters as class variables to extend accessibility to all methods
    header_fill = PatternFill(start_color='CD5C5C', end_color='CD5C5C', fill_type="solid")
    even_row_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type="solid")
    odd_row_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),bottom=Side(style='thick'))
    error_fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')
    alert_fill = PatternFill(start_color='DFFF00', end_color='DFFF00', fill_type='solid')
    improvement_fill = PatternFill(start_color='7CF000', end_color='7CF000', fill_type='solid')
    success_fill = PatternFill(start_color='228B22', end_color='228B22', fill_type='solid')

    # Header & Value : Alignment and Fonts
    value_alignment = Alignment(wrap_text=False, vertical='top', horizontal='left')
    header_alignment = Alignment(wrap_text=False, vertical='top', horizontal='center')
    value_font = Font(name=cell_font_name, size=cell_font_size)
    header_font = Font(name=header_font_name, size=header_font_size, bold=True)

    # Initialize the class with filename as only argument
    def __init__(self, _my_file_name):
        assert _my_file_name.split('.')[-1] in ['xls', 'xlsx'], 'Input file does not have valid excel extension'
        self.my_filename = _my_file_name
        self.my_base_wb = openpyxl.load_workbook(self.my_filename, read_only=False)
        # following line will get the names of worksheets in the workbook
        self.ws_names_in_my_base_wb = self.my_base_wb.sheetnames
        # following line will get the number of worksheets in the workbook
        self.num_ws_in_my_base_wb = len(self.my_base_wb.sheetnames)
        # following line will set the last worksheet in the workbook as active
        self.my_base_active_ws = self.my_base_wb.active
        # following line will get the maximum row number in active worksheet
        self.my_base_active_ws_max_row = self.my_base_wb.active.max_row
        # following line will get the maximum column number in active worksheet
        self.my_base_active_ws_max_col = self.my_base_active_ws.max_column
        # following line will get the column names (as list) in active worksheet
        # Assumption is the column names are in 1st row
        self.my_base_active_ws_col_titles = [_col_name[0].value for _col_name in self.my_base_active_ws.iter_cols()]

    def active_ws(self, _ws_name):
        '''
        Method to set a specific worksheet as active where column names are in row#1
        Argument to this method is: - worksheet name
        '''
        # if the worksheet name exists in the workbook
        if self.if_ws_in_wb(_ws_name):
            # Activate a particular worksheet in a workbook
            # And assign it to the variable in __init__
            # This way, we will be able to use the 'my_base_active_ws' property
            # After object instantiation instead of having a new variable altogether
            self.my_base_active_ws = self.my_base_wb[_ws_name]
            # Expand the properties of the currently active worksheet back to __init__ variables
            self.my_base_active_ws_max_row = self.my_base_active_ws.max_row
            self.my_base_active_ws_max_col = self.my_base_active_ws.max_column
            self.my_base_active_ws_col_titles = [_col_name[0].value for _col_name in self.my_base_active_ws.iter_cols()]
        else:
            print('Worksheet {} not found in workbook'.format(_ws_name))

    ####################################################################################################################
    # Internal Methods for polymorphism support

    def ref_col_idx_name_map(self, _header_row_num) -> dict:
        '''
        Method to get a map of column index and column name
        Argument to this method is: - Row number of header row
        '''
        # Create an empty dictionary
        col_idx_name_map = dict()
        # Iterate over the values of the header row
        # starting with index value of 1
        for _col_idx, _col_cells in enumerate(self.my_base_active_ws.iter_cols(min_row=_header_row_num, max_row=_header_row_num, values_only=True), start=1):
            col_idx_name_map[_col_idx] = _col_cells[0]
        # Return type is dictionary
        return col_idx_name_map

    def ref_col_name_idx_map(self, _header_row_num) -> dict:
        '''
        Method to get a map of column name and column index
        Argument to this method is: - Row number of header row
        '''
        # Create an empty dictionary
        col_name_idx_map = dict()
        # Iterate over the values of the header row
        # starting with index value of 1
        for _col_idx, _col_cells in enumerate(self.my_base_active_ws.iter_cols(min_row=_header_row_num, max_row=_header_row_num, values_only=True), start=1):
            col_name_idx_map[_col_cells[0]] = _col_idx
        # Return type is dictionary
        return col_name_idx_map

    def ref_col_name_letter_map(self, _header_row_num) -> dict:
        '''
        Method to get a map of column name and column letter
        Argument to this method is: - Row number of header row
        '''
        # Create an empty dictionary
        col_name_letter_map = dict()
        # Iterate over the values of the header row
        # starting with index value of 1
        for _col_idx, _col_cells in enumerate(self.my_base_active_ws.iter_cols(min_row=_header_row_num, max_row=_header_row_num, values_only=True), start=1):
            col_name_letter_map[_col_cells[0]] = get_column_letter(_col_idx)
        # Return type is dictionary
        return col_name_letter_map

    def ref_col_letter_name_map(self, _header_row_num) -> dict:
        '''
        Method to get a map of column letter and column map
        Argument to this method is: - Row number of header row
        '''
        col_letter_name_map = dict()
        # Iterate over the values of the header row
        # starting with index value of 1
        for _col_idx, _col_cells in enumerate(self.my_base_active_ws.iter_cols(min_row=_header_row_num, max_row=_header_row_num, values_only=True), start=1):
            col_letter_name_map[(get_column_letter(_col_idx))] = _col_cells[0]
        # Return type is dictionary
        return col_letter_name_map

    def ref_col_letter_idx_map(self, _header_row_num) -> dict:
        '''
        Method to get a map of column letter and column index
        Argument to this method is: - Row number of header row
        '''
        col_letter_idx_map = dict()
        # Iterate over the maximum number of columns (start at 0th index)
        for _col_idx in range(_header_row_num + self.my_base_active_ws.max_column + 1):
            # Note:- For get column letter we need Column Index + 1 since,
            # get column letter does not recognize 0 as first index
            col_letter_idx_map[get_column_letter(_col_idx + 1)] = _col_idx + 1
        # Return type is dictionary
        return col_letter_idx_map

    ####################################################################################################################

    def add_multiple_cols_at_the_end_to_active_ws(self, _header_row_num, _col_names):
        '''
        Method to add multiple columns with names as the last column of active worksheet
        Arguments to this method are:- Column Names (as an iterable), and row number of header row
        '''
        # Iterate over the values in the iterable (column name)
        for _ in _col_names:
            if _ not in self.get_col_names_active_ws(_header_row_num):
                # call the function 'add_new_col_at_the_end_to_active_ws' and pass the column name
                self.add_new_col_at_the_end_to_active_ws(_header_row_num, _)
                # Most important step is to reload the workbook, so that changes from last iteration are reloaded
                self.reload_wb()
            # This method calls save internally, hence its auto save from user perspective
            self.save_wb()

    def add_new_col_at_the_end_to_active_ws(self, _header_row_num, _new_col_name):
        '''
        Method to add a new column with name as the last column of active worksheet
        Arguments to this method are: - New column name and row number of header row
        '''
        # New column will be added only if, it does not exist already
        if _new_col_name not in self.get_col_names_active_ws(_header_row_num):
            # Note:- Since we are adding the column at the end, we need to auto identify its index
            # We are using the existing max_column property of the worksheet and incrementing it by 1
            _new_col_idx = self.my_base_active_ws_max_col + 1
            # Once the columns is added, we correctly identify the cell where the column name will be written
            self.my_base_active_ws.cell(row=_header_row_num, column=_new_col_idx).value = _new_col_name
            # This method calls save internally, hence its auto save from user perspective
        self.save_wb()

    def add_new_col_at_specific_idx_to_active_ws(self, _header_row_num, _new_col_name, _idx_pos):
        '''
        Method to add new column at specific index with a name
        Arguments to the method are: - New Column Name, New Column Position and row number for header
        '''
        # If the new column name already exists, nothing will be done
        if _new_col_name not in self.get_col_names_active_ws(_header_row_num):
            # Insert the column at the specific INDEX position
            # Note: The new column inserted will have a default cell value as None
            self.my_base_active_ws.insert_cols(_idx_pos)
            # Iterate over all the columns of the header row and select the exact cell based on the index position
            # Row number and Column number will provide the exact position
            for _row in self.my_base_active_ws.iter_cols(min_row=_header_row_num,
                                                         max_row=_header_row_num,
                                                         min_col=_idx_pos,
                                                         max_col=_idx_pos):
                for _cell in _row:
                    if _cell.value is None:
                        # Update cell value to new column name
                        _cell.value = _new_col_name
                    # This method calls save internally, hence its auto save from user perspective
            self.save_wb()

    def add_new_ws_at_end_of_wb(self, _new_ws_name):
        '''
        Method to create a new worksheet in workbook
        Argument to this method is: - New worksheet name
        '''
        # New worksheet will be added to workbook, only if it doesn't exist already
        if not self.if_ws_in_wb(_new_ws_name):
            self.my_base_wb.create_sheet(_new_ws_name)
            self.save_wb()

    def add_prefix_to_cell_val(self, _header_row_num, _col_name, _prefix_val):
        '''
        Method to add user provided prefix to value(s) in a given column
        Arguments to this method are:- Row number for header row, Column name & User provided prefix value
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Enumerate over the values from 'get_specific_col_val_by_col_name_in_active_ws' method
        # start with index of (header_row+1) since we need to add prefix values to column values only not column header
        # This updated index start value is used for the row for enumeration
        for _idx, _val in enumerate(self.get_specific_col_val_by_col_name_in_active_ws(_header_row_num, _col_name),
                                    start=_header_row_num + 1):
            # LHS: Assign values to each cell in the column
            # RHS: Concatenate existing value of the cell with the prefixed value
            self.my_base_active_ws.cell(row=_idx, column=_col_idx).value = (
                    str(_prefix_val) + str(self.my_base_active_ws.cell(row=_idx, column=_col_idx).value))
        # Save workbook
        self.save_wb()

    def add_suffix_to_cell_val(self, _header_row_num, _col_name, _suffix_val):
        '''
        Method to add user provided prefix to value in a given column
        Arguments to this method are:- Row number for header row, Column name & User provided prefix value
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Enumerate over the values from 'get_specific_col_val_by_col_name_in_active_ws' method
        # start with index of (header_row+1) since we need to add prefix values to column values only not column header
        # This updated index start value is used for the row for enumeration
        for _idx, _val in enumerate(self.get_specific_col_val_by_col_name_in_active_ws(_header_row_num, _col_name), start=_header_row_num + 1):
            # LHS: Assign values to each cell in the column
            # RHS: Concatenate existing value of the cell with the prefixed value
            self.my_base_active_ws.cell(row=_idx, column=_col_idx).value = (str(self.my_base_active_ws.cell(row=_idx, column=_col_idx).value) + str(_suffix_val))
        # Save workbook
        self.save_wb()

    def adjust_col_width_in_active_ws(self):
        '''
        Method to automatically adjust column width based on max cell value in that columns
        This method has no argument
        '''
        # Enumerate of the columns in the worksheet and start the index from 1
        # Start value is '1' since 'get_column_letter' method does not allow an index of 0
        for _col_num, _col in enumerate(self.my_base_active_ws.columns, start=1):
            # Get the column letter based on the index
            _col_letter = get_column_letter(_col_num)
            # get the length of maximum cell value (width-wise)
            _col_length = max(len(str(cell.value or "")) for cell in _col)
            # create a new variable adjusted width with a new formulated value
            adjusted_width = (_col_length + 2) * 0.95
            if adjusted_width > 30:
                # use the column dimensions parameter and set the width with a new value
                self.my_base_active_ws.column_dimensions[(str(_col_letter))].width = 55
            else:
                # use the column dimensions parameter and set the width with a new value
                self.my_base_active_ws.column_dimensions[(str(_col_letter))].width = adjusted_width
        # Save the workbook after all changes to active worksheet are done.
        # Saving the workbook with every iteration of for loop will increase time by 400%
        self.save_wb()

    def auto_format_active_ws(self, _ws_name, _header_row_num):
        '''
        Method to auto format the rows and columns in the worksheet
        Argument to this method are: - Worksheet name and Row number of header
        '''
        # Activate the worksheet to be formatted
        self.active_ws(_ws_name)
        # Iterate over the number of columns in the worksheet
        for _col in range(1, self.my_base_active_ws_max_col + 1):
            # Iterate over the number of rows in the worksheet
            for _row in range(_header_row_num, self.my_base_active_ws_max_row + 1):
                # Set font, border and alignment for non header cells
                # Identify cell range for non-header row
                filled_cells = self.my_base_active_ws.cell(_row, _col)
                # set the alignment of the non-header row
                filled_cells.alignment = self.value_alignment
                # set the font of the non-header row
                filled_cells.font = self.value_font
                # set the border for non-header row
                filled_cells.border = self.thin_border
            # Set font, border and alignment for header cells
            # Identify cell range for header row
            _cell_header = self.my_base_active_ws.cell(_header_row_num, _col)
            # set the alignment of the header row
            _cell_header.alignment = self.header_alignment
            # set the font of the header row
            _cell_header.font = self.header_font
            # set the color fill for header row
            _cell_header.fill = self.header_fill
            # set the border for header row
            _cell_header.border = self.thick_border
        self.save_wb()

    def copy_and_paste_column_values_by_col_name(self, _src_col_name, _tgt_col_name, _header_row_num):
        '''
        Method to copy and paste data from one column to another, excluding header
        Arguments to this method are: - name of source column, name of target column & row number for header
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _src_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_src_col_name])
        _tgt_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_tgt_col_name])
        # Enumerate over values (by rows) of the worksheet
        for _xl_row_idx, _row_val in enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1):
            # Exclude the header row
            if _xl_row_idx != _header_row_num:
                # Write the contents of cells to target column from source column
                self.my_base_active_ws.cell(row=_xl_row_idx,
                                            column=_tgt_col_idx).value = self.my_base_active_ws.cell(
                    row=_xl_row_idx, column=_src_col_idx).value
        # Save workbook
        self.save_wb()

    def change_cell_color_by_cell_value_near_match(self, _header_row_num, _col_name, _cell_value, _fill_color):
        '''
        Method to change row color based on a near match of cell value in particular column
        Arguments to this method: - Row number for header row, Column name, value of cell, user selected fill color
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Get the list of indexes where near match of searched value if found excluding the header row
        _row_idx = self.get_row_idx_lst_based_on_search_val_specific_col_near_match(_header_row_num, _col_name,
                                                                                    _cell_value)
        # Set the color fill based on user selected color
        _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                                fill_type='solid')
        # Iterate over the list of row indexes identified earlier
        for _row_num in _row_idx:
            # Identify cell range with matching criteria
            # Note: The updated index value is to be used since we are writing to the cell
            filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_col_idx)
            # set the color fill for row
            filled_cells.fill = _my_color
        self.save_wb()

    def change_cell_color_by_cell_value_exact_match(self, _header_row_num, _col_name, _cell_value, _fill_color):
        '''
        Method to change row color based on an exact match of cell value in particular column
        Arguments to this method: - Row number for header row, Column name, value of cell, user selected fill color
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Get the list of indexes where near match of searched value if found excluding the header row
        _row_idx = self.get_row_idx_lst_based_on_search_val_specific_col_exact_match(_header_row_num, _col_name,
                                                                                     _cell_value)
        # Set the color fill based on user selected color
        _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                                fill_type='solid')
        # Iterate over the list of row indexes identified earlier
        for _row_num in _row_idx:
            # Identify cell range with matching criteria
            # Note: The updated index value is to be used since we are writing to the cell
            filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_col_idx)
            # set the color fill for row
            filled_cells.fill = _my_color
        self.save_wb()

    def change_row_color_empty_cell_in_col(self, _header_row_num, _col_name, _fill_color):
        '''
        Method to change row color based on an empty cell in particular column
        Arguments to this method: - Row number for header row, Column name, and user selected fill color
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Iterate over the values of the column and get row indexes of empty cells excluding header row
        # The Excel columns start with 1, however when iterating, the tuples start with index 0
        _row_idx = [_xl_row_idx for _xl_row_idx, _row_val in
                    enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1) if not _row_val[_col_idx - 1]
                    if
                    _xl_row_idx != _header_row_num]
        # define the color pattern and using the class variable match the color hex code
        _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                                fill_type='solid')
        # Iterate over the list of row indexes
        for _row_num in _row_idx:
            # Iterate over the columns of the row
            for _ in range(1, self.my_base_active_ws_max_col + 1):
                # Identify the cells based on row and column indexes
                filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_)
                # apply color
                filled_cells.fill = _my_color
        self.save_wb()

    def change_row_color_non_empty_cell_in_col(self, _header_row_num, _col_name, _fill_color):
        '''
        Method to change row color based on a non-empty cell in particular column
        Arguments to this method: - Row number for header row, Column name, and user selected fill color
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Iterate over the values of the column and get row indexes of non-empty cells excluding header row
        # The Excel columns start with 1, however when iterating, the tuples start with index 0
        _row_idx = [_xl_row_idx for _xl_row_idx, _row_val in
                    enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1) if _row_val[_col_idx - 1] if
                    _xl_row_idx != _header_row_num]
        # define the color pattern and using the class variable match the color hex code
        _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                                fill_type='solid')
        # Iterate over the list of row indexes
        for _row_num in _row_idx:
            # Iterate over the columns of the row
            for _ in range(1, self.my_base_active_ws_max_col + 1):
                # Identify the cells based on row and column indexes
                filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_)
                # apply color
                filled_cells.fill = _my_color
        self.save_wb()

    def change_row_color_by_cell_value_near_match(self, _header_row_num, _col_name, _cell_value, _fill_color):
        '''
        Method to change row color based on a near match of cell value in particular column
        Arguments to this method: - Row number for header row, Column name, value of cell, user selected fill color
        '''
        # Get the list of indexes where near match of searched value if found excluding the header row
        _row_idx = self.get_row_idx_lst_based_on_search_val_specific_col_near_match(_header_row_num, _col_name,
                                                                                    _cell_value)
        # Set the color fill based on user selected color
        _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                                fill_type='solid')
        # Iterate over the list of row indexes identified earlier
        for _row_num in _row_idx:
            # Iterate over each column of the row
            for _ in range(1, self.my_base_active_ws_max_col + 1):
                # Identify cell range with matching criteria
                filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_)
                # set the color fill for row
                filled_cells.fill = _my_color
        self.save_wb()

    def change_row_color_by_cell_value_exact_match(self, _header_row_num, _col_name, _cell_value, _fill_color):
        '''
        Method to change row color based on an exact match of cell value in particular column
        Arguments to this method: - Row number for header row, Column name, value of cell, user selected fill color
        '''
        # Get the list of indexes where near match of searched value if found excluding the header row
        _row_idx = self.get_row_idx_lst_based_on_search_val_specific_col_exact_match(_header_row_num, _col_name,
                                                                                     _cell_value)
        # Set the color fill based on user selected color
        _my_color = PatternFill(start_color=self._my_color_map[_fill_color], end_color=self._my_color_map[_fill_color],
                                fill_type='solid')
        # Iterate over the list of row indexes identified earlier
        for _row_num in _row_idx:
            # Iterate over each column of the row
            for _ in range(1, self.my_base_active_ws_max_col + 1):
                # Identify cell range with matching criteria
                filled_cells = self.my_base_active_ws.cell(row=_row_num, column=_)
                # set the color fill for row
                filled_cells.fill = _my_color
        self.save_wb()

    def copy_src_col_paste_tgt_col_diff_sheets(self, _header_row_num, _src_ws_name, _src_col_name, _tgt_ws_name, _tgt_col_name):
        '''
        Method to copy data from source and paste it to target of different worksheets in same workbook
        Arguments to this method are: - Source worksheet name, Column Letter of Source, Target worksheet name
        and Column letter of Target
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        # Since we are dealing with multiple worksheets in this case
        # We must activate each worksheet before pulling column indexes
        self.active_ws(_src_ws_name)
        _src_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_src_col_name])
        self.active_ws(_tgt_ws_name)
        _tgt_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_tgt_col_name])
        # Enumerate over values (by rows) of the worksheet
        for _xl_row_idx, _row_val in enumerate(self.my_base_wb[_src_ws_name].iter_rows(values_only=True), start=1):
            # Write the contents of cells to target column from source column excluding header row
            self.my_base_wb[_tgt_ws_name].cell(row=_xl_row_idx + _header_row_num, column=_tgt_col_idx).value = \
                self.my_base_wb[_src_ws_name].cell(row=_xl_row_idx + _header_row_num, column=_src_col_idx).value
        # Save workbook
        self.save_wb()

    def copy_ws_src_to_tgt(self, _src_ws, _tgt_ws):
        '''
        Method to copy data from source worksheet to target worksheet
        Arguments to this method are: -  source worksheet name and target worksheet name
        Note: This method will create a new worksheet, if the target worksheet does not exist
        '''
        # Code block will only proceed if source worksheet exists
        if self.if_ws_in_wb(_src_ws):
            # If the target worksheet exists, code will proceed to iterations and copy
            if self.if_ws_in_wb(_tgt_ws):
                pass
            # If the target worksheet does not exist, new worksheet will be created
            # Name of newly created worksheet will the 2nd argument
            else:
                self.add_new_ws_at_end_of_wb(_tgt_ws)
                # Once the target worksheet is added, save the workbook
                self.save_wb()
            # Special Note: If we use iterrows() or itercolumns() the cell value assignment fails
            # Since the active keyword is not used for sheet activation
            # iterate over the number of columns  in source
            for _row in range(1, self.my_base_wb[_src_ws].max_row + 1):
                # Iterate over the number of rows in source
                for _col in range(1, self.my_base_wb[_src_ws].max_column + 1):
                    # Assign the values to each target cell from source cell
                    self.my_base_wb[_tgt_ws].cell(row=_row, column=_col).value = self.my_base_wb[_src_ws].cell(row=_row,
                                                                                                               column=_col).value
            self.save_wb()

    def copy_header_row_src_ws_to_tgt_ws(self, _header_row_num, _src_ws_name, _tgt_ws_name):
        '''
        Method to copy the header value from one worksheet to another
        Arguments to this method are: Row number of the header row, source worksheet name and target worksheet name  
        '''
        # Capture the column names from the source worksheet names using list comprehension
        _src_col_name = [_ for _ in self.my_base_wb[_src_ws_name].iter_rows(min_row=_header_row_num, max_row=(_header_row_num), values_only=True)][0]
        # Iterate over the column names and the column indexes starting at position = 1
        for _col_idx, _col_val in enumerate(_src_col_name, start=1):
            # Write the column name to the target worksheet based on the iterated index and value
            self.my_base_wb[_tgt_ws_name].cell(row=_header_row_num, column=_col_idx).value = _col_val
        # Save the workbook
        self.save_wb()

    def convert_multi_col_val_to_single_col_in_active_ws(self):
        '''
        Method to convert multi-column values in to single column values from active worksheet
        '''
        # Return a list of values
        return [_cell.value for _ in self.my_base_active_ws.columns for _cell in _]

    def create_dict_from_values(self):
        '''
        Create a dictionary from Excel cell values
        No arguments to this method
        '''
        # Create an empty dictionary
        _my_dict = dict()
        # get the unique keys (numeric) from the first column of the worksheet
        _my_keys = set([_row[0] for _row in self.my_base_active_ws.values if str(_row[0]).isdigit()])
        # Iterate over the keys and add the values as a list
        for _ in _my_keys:
            # using list comprehension add the values to the key or keys
            _my_dict[_] = [_row[1] for _row in self.my_base_active_ws.values if _row[0] == _]
        return _my_dict

    def df_to_rows(self, _dataframe, _ws_name):
        '''
        Method to convert dataframe to excel rows
        Arguments to this method are: - Dataframe and Worksheet name
        '''
        # Iterate over the rows in dataframe
        for _ in dataframe_to_rows(_dataframe, index=False):
            # Append the rows to worksheet
            self.my_base_wb[_ws_name].append(_)
        # Save workbook
        self.save_wb()

    # Method to delete a row based on row index
    def del_row_by_idx(self, _row_idx):
        return self.my_base_active_ws.delete_rows(_row_idx)

    def del_row_by_val_in_col_exact_match(self, _header_row_num, _col_name, _cell_val):
        '''
        Method to delete rows based on an exact match of cell value in a given column
        Argument to this method are: - Row number of the header row, Column name and cell value
        '''
        # Check if provided column name exists in the worksheet using 'ref_col_name_idx_map' method
        # 'ref_col_name_idx_map' method has return type of dict
        if _col_name in self.ref_col_name_idx_map(_header_row_num).keys():
            # Check if the value provided has an exact match in the given column
            if self.get_row_idx_lst_based_on_search_val_specific_col_exact_match(_header_row_num, _col_name,
                                                                                 _cell_val):
                # Get the list of indexes based on match condition
                _depleting_idx = [_ for _ in
                                  self.get_row_idx_lst_based_on_search_val_specific_col_exact_match(_header_row_num,
                                                                                                    _col_name,
                                                                                                    _cell_val)]
                # Iterate for all values in _depleting_idx
                while _depleting_idx:
                    # Delete the row for the iterated index value
                    # Start from the last index value since it will not change position of index, post delete
                    # Deletion of row from top will move the following row one row up there by changing the index
                    # Causing incorrect row to be deleted starting from 2nd iteration
                    self.my_base_active_ws.delete_rows(_depleting_idx[-1])
                    # Save the workbook
                    self.save_wb()
                    # delete the last element from the index to deplete the iteration count
                    del _depleting_idx[-1]

    def del_row_by_val_in_col_near_match(self, _header_row_num, _col_name, _search_val):
        '''
        Method to delete rows based on a near match of cell value in a given column
        Argument to this method are: - Row number of the header row, Column name and cell value
        '''
        # Check if provided column name exists in the worksheet using 'ref_col_name_idx_map' method
        # 'ref_col_name_idx_map' method has return type of dict
        if _col_name in self.ref_col_name_idx_map(_header_row_num).keys():
            # Check if the value provided has a near match in the given column
            if self.get_row_idx_lst_based_on_search_val_specific_col_near_match(_header_row_num, _col_name,
                                                                                _search_val):
                # Get the list of indexes based on match condition
                _depleting_idx = [_ for _ in
                                  self.get_row_idx_lst_based_on_search_val_specific_col_near_match(_header_row_num,
                                                                                                   _col_name,
                                                                                                   _search_val)]
                # Iterate for all values in _depleting_idx
                while _depleting_idx:
                    # Delete the row for the iterated index value
                    # Start from the last index value since it will not change position of index, post delete
                    # Deletion of row from top will move the following row one row up there by changing the index
                    # Causing incorrect row to be deleted starting from 2nd iteration
                    self.my_base_active_ws.delete_rows(_depleting_idx[-1])
                    # Save the workbook
                    self.save_wb()
                    # delete the last element from the index to deplete the iteration count
                    del _depleting_idx[-1]

    def drop_col_by_name_in_active_ws(self, _header_row_num, _ws_name, _col_name):
        '''
        Method to drop column by name
        Arguments to this method are:- Row number for header row, worksheet name and column name
        '''
        # Activate the worksheet
        self.active_ws(_ws_name)
        # get the column index based on the column name
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Delete the column using the column index
        self.my_base_active_ws.delete_cols(_col_idx, 1)
        # Save the workbook
        self.save_wb()

    def drop_col_by_letter_in_active_ws(self, _header_row_num, _ws_name, _col_letter):
        '''
        Method to drop the column by column letter
        Arguments to this method are:- Row number for header row, worksheet name and column letter
        '''
        # Activate the worksheet
        self.active_ws(_ws_name)
        # get the column index based on the column name
        _col_idx = column_index_from_string(_col_letter)
        # Delete the column using the column index
        self.my_base_active_ws.delete_cols(_col_idx, 1)
        # Save the workbook
        self.save_wb()

    def filter_rows_by_val_in_any_col_in_active_ws(self, _header_row_num, _filter_val):
        '''
        Method to filter a row based on a value present in any column of worksheet
        Argument to this method is:- value to be used as filter
        '''
        # Create an empty list to hold values
        # One can always use list comprehension if needed
        returned_val = []
        # Iterate over the value in the worksheet(column and rows)
        # Note: We are directly iterating over values instead of cell references
        for _main_rec in self.my_base_active_ws.iter_rows(min_row=_header_row_num + 1, values_only=True):
            # If the searched value is found in any of the cells of any column
            if str(_filter_val).casefold() in str(_main_rec).casefold():
                # Append the corresponding row to the empty list
                returned_val.append(_main_rec)
        return returned_val

    def filter_rows_by_val_not_in_any_col_in_active_ws(self, _header_row_num, _filter_val):
        '''
        Method to filter a row based on a value present in any column of worksheet
        Argument to this method is:- value to be used as filter
        '''
        # Create an empty list to hold values
        # One can always use list comprehension if needed
        returned_val = []
        # Iterate over the value in the worksheet(column and rows)
        # Note: We are directly iterating over values instead of cell references
        for _main_rec in self.my_base_active_ws.iter_rows(min_row=_header_row_num + 1, values_only=True):
            # If the searched value is found in any of the cells of any column
            if str(_filter_val).casefold() not in str(_main_rec).casefold():
                # Append the corresponding row to the empty list
                returned_val.append(_main_rec)
        return returned_val

    def filter_rows_by_val_in_a_col_in_active_ws(self, _header_row_num, _filter_col_name, _filter_val):
        '''
        Method to filter rows based on a value present in a specific column of worksheet
        Arguments to this method are: - value used as filter, column idx where filter will be applied
        '''
        # Create an empty list to hold values
        # One can always use list comprehension if needed
        returned_val = []
        _filter_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_filter_col_name])
        # Iterate over the value in the worksheet(column and rows)
        # Note: We are directly iterating over values instead of cell references
        for _main_rec in self.my_base_active_ws.iter_rows(min_row=_header_row_num + 1, values_only=True):
            # If the searched value is found in any cell of the specific column
            # The Excel columns start with 1, however when iterating, the tuples start with index 0
            if str(_filter_val).casefold() == str(_main_rec[_filter_col_idx - 1]).casefold():
                # Append the corresponding row to the empty list
                returned_val.append(_main_rec)
        return returned_val

    def filter_rows_by_val_not_in_a_col_in_active_ws(self, _header_row_num, _filter_col_name, _filter_val):
        '''
        Method to filter rows based on a value present in a specific column of worksheet
        Arguments to this method are: - value used as filter, column idx where filter will be applied
        '''
        # Create an empty list to hold values
        # One can always use list comprehension if needed
        returned_val = []
        _filter_col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_filter_col_name])
        # Iterate over the value in the worksheet(column and rows)
        # Note: We are directly iterating over values instead of cell references
        for _main_rec in self.my_base_active_ws.iter_rows(min_row=_header_row_num + 1, values_only=True):
            # If the searched value is found in any cell of the specific column
            # The Excel columns start with 1, however when iterating, the tuples start with index 0
            if str(_filter_val).casefold() != str(_main_rec[_filter_col_idx - 1]).casefold():
                # Append the corresponding row to the empty list
                returned_val.append(_main_rec)
        return returned_val

    def format_header_row(self, _header_row_num, _cell_color, _font_color):
        '''
        Method to format header row with user defined, cell color and font color
        Arguments to this method are: - Row number of the header row, fill color for the cells in header row
        Font color in the cells
        '''
        # Iterate over the cells in the header row to the max column size
        for _col_idx in range(1, self.my_base_active_ws_max_col + 1):
            # Set the header cells for easy style formatting
            header_cells = self.my_base_active_ws.cell(row=_header_row_num, column=_col_idx)
            # Set header cell alignment
            header_cells.alignment = Alignment(wrap_text=False, vertical='top', horizontal='center')
            # Set header cell border
            header_cells.border = self.thick_border
            # Fill the cells with user selected color from the color map
            header_cells.fill = PatternFill(start_color=self._my_color_map[_cell_color],
                                            end_color=self._my_color_map[_cell_color],
                                            fill_type="solid")
            # Change the color of the font in cells with user selected color
            header_cells.font = Font(bold=True, size=self.header_font_size,
                                     name=self.header_font_name,
                                     color=self._my_color_map[_font_color])
        self.save_wb()

    def format_val_in_column(self, _header_row_num, _col_name, _h_align, _font_color):
        '''
        Method to format columns values with user defined, horizontal alignment ant font color
        Arguments to this method are: - Row number of the header row, column name,
        Font color in the cells
        '''
        note = 'Valid Values are:- right, justify, general, centerContinuous, fill, center, left, distributed'
        assert _h_align in ['right', 'justify', 'general', 'centerContinuous', 'fill', 'center', 'left',
                            'distributed'], note
        # Valid values for _h_align are (â€˜rightâ€™, â€˜justifyâ€™, â€˜generalâ€™,
        # â€˜centerContinuousâ€™, â€˜fillâ€™, â€˜centerâ€™, â€˜leftâ€™, â€˜distributedâ€™)
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Iterate over the rows in the given column until last row excluding header row
        for _row_idx in range(_header_row_num + 1, self.my_base_active_ws_max_row + 1):
            # Set the value cells for easy style formatting
            # The Excel columns start with 1, however when iterating, the tuples start with index 0
            value_cells = self.my_base_active_ws.cell(row=_row_idx, column=_col_idx)
            # Set value cell alignment
            value_cells.alignment = Alignment(wrap_text=False, vertical='top', horizontal=_h_align)
            # Set value cell border
            value_cells.border = self.thin_border
            # Change the color of the font in cells with user selected color
            value_cells.font = Font(bold=False, size=self.cell_font_size,
                                    name=self.cell_font_name,
                                    color=self._my_color_map[_font_color])
        self.save_wb()

    def get_col_names_active_ws(self, _header_row_num):
        '''
        Method to get column names of the active worksheet
        '''
        _col_names = [_ for _ in self.ref_col_idx_name_map(_header_row_num).values()]
        # Return type is a list
        return _col_names

    def get_cell_coordinate_by_search_string(self, _search_val):
        '''
        Method to get list of cell coordinate by a search value
        Argument to this method is:- search string
        '''
        # Create an empty list
        _match_row_idx_lst = []
        # Enumerate over the values (in rows) with starting index of 1
        for _xl_row_idx, _row_val in enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1):
            # If search string is in the cell value
            # Ignore Case while checking membership
            if str(_search_val).casefold() in str(_row_val).casefold():
                # Iterate over the length of cell values to get the index
                for _idx in range(len(_row_val)):
                    if str(_search_val).casefold() == str(_row_val[_idx]).casefold():
                        # get the index position and concatenate the Column Letter with row index
                        # Append the identified cell coordinate to empty list
                        _match_row_idx_lst.append(str(get_column_letter(_idx + 1) + str(_xl_row_idx)))
        # Return type is a list
        return _match_row_idx_lst

    def get_specific_col_val_by_col_name_in_active_ws(self, _header_row_num, _col_name):
        '''
        Method to get column values for a given column name
        Arguments to this method are:- Row number for header row & column name
        '''
        # Check if provided column name exists in the worksheet
        if _col_name in self.ref_col_idx_name_map(_header_row_num).values():
            # Fetch the values from the column provided
            # Skip the header row
            # Fetch the column index from column name using 'ref_col_name_letter_map' method
            return [_col_value[0] for _col_value in self.my_base_active_ws.iter_rows(min_row=_header_row_num + 1,
                                                                                     min_col=column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name]),
                                                                                     max_col=column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name]),
                                                                                     values_only=True)]

    def get_specific_col_val_by_col_idx_in_active_ws(self, _header_row_num, _col_idx):
        '''
        Method to get column values for a given column index
        Arguments to this method are:- Row number for header row & column index
        Note:- Column index starts with 1
        '''
        # Fetch the values from the column index provided
        # Skip the header row
        assert _col_idx >= 1, 'ValueError: Row or column values must be at least 1'
        try:
            return [_col_value[0] for _col_value in self.my_base_active_ws.iter_rows(min_row=_header_row_num + 1,
                                                                                     min_col=_col_idx,
                                                                                     max_col=_col_idx,
                                                                                     values_only=True)]
        except ValueError:
            print('ValueError: Row or column values must be at least 1')

    def get_specific_col_val_by_col_letter_in_active_ws(self, _header_row_num, _col_letter):
        '''
        Method to get column values for a given column letter
        Arguments to this method are:- Row number for header row & column letter
        '''
        assert _col_letter in self.ref_col_letter_name_map(
            _header_row_num).keys(), 'Column Letter Outside Data-Set Range'
        _col_idx = column_index_from_string(_col_letter)
        try:
            return [_col_value[0] for _col_value in self.my_base_active_ws.iter_rows(min_row=_header_row_num + 1,
                                                                                     min_col=_col_idx,
                                                                                     max_col=_col_idx,
                                                                                     values_only=True)]
        except AssertionError:
            print('AssertionError: Column Letter Outside Data-Set Range')

    def get_row_idx_lst_based_on_search_val_specific_col_near_match(self, _header_row_num, _col_name, _search_val):
        '''
        Method to get row indexes of searched values in a specific column with near match condition
        Arguments to this method are:- Row number for header row, Column name & Search value
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Get the list of indexes where near match of searched value is found excluding the header row
        # The Excel columns start with 1, however when iterating, the tuples start with index 0
        _row_idx_list = [_xl_row_idx for _xl_row_idx, _row_val in
                         enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1) if
                         str(_search_val) in str(_row_val[_col_idx - 1]) if _xl_row_idx != _header_row_num]
        # Return type is list
        return _row_idx_list

    def get_row_idx_lst_based_on_search_val_specific_col_exact_match(self, _header_row_num, _col_name, _search_val):
        '''
        Method to get row indexes of searched values in a specific column with exact match condition
        Arguments to this method are:- Row number for header row, Column name & Search value
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Get the list of indexes where exact match of searched value is found excluding the header row
        # Example of list comprehension with nested if conditions
        # The Excel columns start with 1, however when iterating, the tuples start with index 0
        _row_idx_list = [_xl_row_idx for _xl_row_idx, _row_val in
                         enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1) if
                         str(_search_val).casefold() in str(_row_val[_col_idx - 1]).casefold() if
                         str(_row_val[_col_idx - 1]).casefold() == str(_search_val).casefold() if
                         _xl_row_idx != _header_row_num]
        # Return type is list
        return _row_idx_list

    def get_row_idx_lst_of_empty_rows_specific_col(self, _header_row_num, _col_name):
        '''
        Method to get row indexes of empty rows in a specific column
        Arguments to this method are: - Row number of header row and Column name
        '''
        # Fetch the column index from column name using 'ref_col_name_idx_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # Get the list of indexes of empty rows excluding the header row
        _row_idx_list = [_xl_row_idx for _xl_row_idx, _row_val in
                         enumerate(self.my_base_active_ws.iter_rows(values_only=True), start=1) if
                         not _row_val[_col_idx - 1]]
        # Return type is list
        return _row_idx_list

    def get_specific_row_val_as_list_in_active_ws(self, _val_row_num):
        '''
        Method to get values for specific row in a given worksheet
        Argument to this method is: - Row number of values to be fetched
        '''
        for _col in self.my_base_active_ws.iter_cols():
            # Iterate once for the specific row number in active worksheet
            for _row in self.my_base_active_ws.iter_rows(min_row=_val_row_num, max_row=_val_row_num):
                # Return a list of values
                return [_cell.value for _cell in _row]

    def get_specific_row_val_as_str_in_active_ws(self, _val_row_num):
        '''
        Method to get values for specific row in a given worksheet
        Argument to this method is: - Row number of values to be fetched
        '''
        # Iterate once for all the columns in the active worksheet
        for _col in self.my_base_active_ws.iter_cols(min_col=1, max_col=1):
            # Iterate once for the specific row number in active worksheet
            for _row in self.my_base_active_ws.iter_rows(min_row=_val_row_num, max_row=_val_row_num):
                # Return a string of values
                return ', '.join(map(str, [_cell.value for _cell in _row]))

    def get_specific_row_val_as_dict_in_active_ws(self, _header_row_num, _val_row_num):
        '''
        Method to get values for specific row in a given worksheet
        Arguments to this method are: - Row number of values to be fetched and Row number of header row
        '''
        # Iterate once for all the columns in the active worksheet
        _col_keys = [_ for _ in self.ref_col_name_idx_map(_header_row_num).keys()]
        # Create an empty list
        _row_values = []
        # Iterate once for all the columns in the active worksheet
        for _col in self.my_base_active_ws.iter_cols():
            # Iterate once for the specific row number in active worksheet
            for _row in self.my_base_active_ws.iter_rows(min_row=_val_row_num, max_row=_val_row_num):
                # for each cell in the row
                for _cell in _row:
                    # Return a list of values
                    _row_values.append(_cell.value)
            # Using the  _col_keys  and _row_values create a dictionary
            _result_set = dict(zip(_col_keys, _row_values))
            # Return a dictionary of values
            return _result_set

    def if_ws_in_wb(self, _ws_name):
        '''
        Method to check if a given worksheet exists in workbook
        Argument to this method is: workbook name
        '''
        # if worksheet exists in list of worksheets of workbook returns True else False
        if _ws_name in self.ws_names_in_my_base_wb:
            return True
        return False

    def read_values_from_cell_range_active_ws(self, _header_row_num, _start_range, _end_range):
        '''
        Method to read values from a cell range
        Arguments to this method are:- Row number for header row, Start range and Ending range
        '''
        # Split start range to determine 'Column Letter' and Row index
        _start_rng_col = ''.join([_ for _ in _start_range if _.isalpha()])
        _start_rng_row = ''.join([_ for _ in _start_range if _.isdigit()])
        _end_rng_col = ''.join([_ for _ in _end_range if _.isalpha()])
        _end_rng_row = ''.join([_ for _ in _end_range if _.isdigit()])
        # Create an empty list to hold iterated values
        _my_val = []
        # Iterate over the values (row-wise) of the active worksheet
        # minimum row - Is the row number from the start range
        # maximum row - Is the row number from the end range
        # minimum column - Is the column index fetched from 'ref_col_letter_idx_map' method using column letter
        # maximum column - Is the column index fetched from 'ref_col_letter_idx_map' method using column letter + 1
        # 1 is added to maximum column to make the last column inclusive for result return
        for _val in self.my_base_active_ws.iter_rows(values_only=True,
                                                     min_row=int(_start_rng_row),
                                                     min_col=int(self.ref_col_letter_idx_map(_header_row_num)[_start_rng_col]),
                                                     max_row=int(_end_rng_row),
                                                     max_col=int(self.ref_col_letter_idx_map(_header_row_num)[_end_rng_col] + 1)):
            _my_val.append(_val)
        return _my_val

    def remove_ws_from_wb(self, _ws_name):
        '''
        Method to remove a worksheet from a workbook
        Arguments to this method is: - worksheet name
        '''
        try:
            del self.my_base_wb[_ws_name]
            self.save_wb()
            return '{} : worksheet removed successfully'.format(_ws_name)
        except KeyError:
            print('Worksheet by name:- "{}" was not found'.format(_ws_name))

    def sort_asc_row_val_in_col_sep_by_delim_active_ws(self, _header_row_num, _col_name, _delim, _join_by_delim):
        '''
        Method to sort values in a given row in an iterative manner
        Arguments to this method are: - Index of column, row number of header, delimiter the values are separated by
        and delimiter value by which the values will be joined by
        '''
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        # The Excel columns start with 1, however when iterating, the tuples start with index 0
        for _idx, _val in enumerate(self.get_specific_col_val_by_col_name_in_active_ws(_header_row_num, _col_name), start=1):
            # Bug fix to handle None values in the column
            # Otherwise it will fail during the sorting for split method
            if _val is not None:
                self.my_base_active_ws.cell(row=_idx + _header_row_num, column=_col_idx).value = _join_by_delim.join(sorted([_ for _ in _val.split(_delim)]))
        self.save_wb()

    def sort_dsc_row_val_in_col_sep_by_delim_active_ws(self, _header_row_num, _col_name, _delim, _join_by_delim):
        '''
        Method to sort values in a given row in an iterative manner
        Arguments to this method are: - Index of column, row number of header, delimiter the values are separated by
        and delimiter value by which the values will be joined by
        '''
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        for _idx, _val in enumerate(self.get_specific_col_val_by_col_name_in_active_ws(_header_row_num, _col_name),
                                    start=1):
            # Bug fix to handle None values in the column
            # Otherwise it will fail during the sorting for split method
            if _val is not None:
                self.my_base_active_ws.cell(row=_idx + _header_row_num,
                                            column=_col_idx).value = _join_by_delim.join(
                    sorted([_ for _ in _val.split(_delim)], reverse=True))
        self.save_wb()

    def reload_wb(self):
        '''
        Method to reload the workbook after any value update or column additions
        No argument to this method
        '''
        # Call the initialization method with the file name to reload
        # since "openpyxl.load_workbook(self.my_filename, read_only=False)" is part of initialization
        self.__init__(self.my_filename)

    def save_wb(self):
        '''
        Method to save the workbook
        No arguments to this method
        '''
        return self.my_base_wb.save(self.my_filename)

    def save_wb_as_new_file(self, _new_file_name):
        '''
        Method to save the workbook to new file
        Argument for this method is: - New File Name
        '''
        assert _new_file_name.split('.')[-1] in ['xls', 'xlsx'], 'Input file does not have valid excel extension'
        return self.my_base_wb.save(_new_file_name)

    def write_values_to_empty_cells_in_column(self, _header_row_num, _col_name, _write_val):
        '''
        Method to write static value to empty cells in a given column
        Arguments to this method are:- Row number of the header wow, Column name and static to be written
        '''
        # Fetch the column index from column name using 'ref_col_name_letter_map' method
        _col_idx = column_index_from_string(self.ref_col_name_letter_map(_header_row_num)[_col_name])
        for _row_idx in self.get_row_idx_lst_of_empty_rows_specific_col(_header_row_num, _col_name):
            # write user defined value to the cells at specific row indexes
            self.my_base_active_ws.cell(row=_row_idx, column=_col_idx).value = _write_val
        # save workbook
        self.save_wb()